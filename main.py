import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def main():
    st.title('Aplicação de Análise de Arquivos XLSX')
    st.write("Faça o upload dos dois arquivos XLSX para começar a análise.")

    # Upload dos dois arquivos
    file1 = st.file_uploader("Upload da base de dados XLSX", type="xlsx")
    file2 = st.file_uploader("Upload do arquivo para validar XLSX", type="xlsx")

    if file1 is not None and file2 is not None:
        if st.button("Executar Comparação"):
            try:
                # Leitura dos arquivos XLSX
                base_dados = pd.read_excel(file1)
                arquivo_validar = pd.read_excel(file2)

                # Normalizar nomes das colunas removendo espaços extras
                base_dados.columns = base_dados.columns.str.strip()
                arquivo_validar.columns = arquivo_validar.columns.str.strip()

                st.write("Visualização do Primeiro Arquivo:")
                st.dataframe(base_dados.head())
                st.write("Visualização do Segundo Arquivo:")
                st.dataframe(arquivo_validar.head())

                if set(base_dados.columns) != set(arquivo_validar.columns):
                    st.warning("As planilhas têm colunas diferentes. Certifique-se de que ambas têm as mesmas colunas.")
                else:
                    # Criar um DataFrame booleano para armazenar quais linhas têm diferenças (ignorando cabeçalho)
                    diferencas = arquivo_validar.iloc[1:].apply(lambda row: any(
                        str(row[col]).lower() not in base_dados[col].astype(str).str.lower().values for col in
                        arquivo_validar.columns if col in row), axis=1)

                    # Carregar o arquivo original para edição
                    file2.seek(0)  # Resetar ponteiro para garantir leitura correta
                    wb = load_workbook(file2)
                    ws = wb.active

                    fill_yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Para linhas
                    fill_blue = PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type="solid")  # Para colunas

                    # Pintar as linhas e colunas com diferenças (ignorando cabeçalho)
                    for i, row in enumerate(arquivo_validar.iloc[1:].itertuples(index=False, name=None),
                                            start=3):  # Começa na linha 3 para ignorar cabeçalho
                        has_difference = False
                        for j, col_name in enumerate(arquivo_validar.columns, start=1):
                            if str(row[j - 1]).lower() not in base_dados[col_name].astype(str).str.lower().values:
                                has_difference = True

                        if has_difference:
                            for cell in ws[i]:
                                cell.fill = fill_yellow  # Destaca toda a linha primeiro

                            # Agora, destacar as células específicas em azul
                            for j, col_name in enumerate(arquivo_validar.columns, start=1):
                                if str(row[j - 1]).lower() not in base_dados[col_name].astype(str).str.lower().values:
                                    ws.cell(row=i, column=j).fill = fill_blue

                    # Salvar o arquivo modificado
                    output = BytesIO()
                    wb.save(output)
                    output.seek(0)

                    st.download_button(
                        label="Baixar arquivo com diferenças destacadas",
                        data=output,
                        file_name="planilha_destacada.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                    st.success(
                        "Arquivo gerado com sucesso! As diferenças foram destacadas: linhas em amarelo e colunas em azul.")
            except Exception as e:
                st.error(f"Erro ao processar os arquivos: {e}")

if __name__ == "__main__":
    main()
